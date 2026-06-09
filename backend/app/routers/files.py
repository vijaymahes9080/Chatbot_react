import time
from typing import List
from datetime import datetime
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, status
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.file import Document
from app.models.user import User
from app.schemas.file import DocumentResponse, FileUploadResponse
from app.services.auth import get_current_user
from app.services.rag import RAGPipeline

router = APIRouter(prefix="/files", tags=["files"])
rag_pipeline = RAGPipeline()

@router.get("", response_model=List[DocumentResponse])
def list_documents(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return db.query(Document).filter(Document.user_id == current_user.id).all()

@router.post("/upload", response_model=FileUploadResponse)
async def upload_document(
    file: UploadFile = File(...),
    chat_id: str = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Uploads a document, extracts text based on file format (PDF, DOCX, TXT, CSV, Excel),
    and indexes the document chunks inside the vector database.
    """
    contents = await file.read()
    file_size_mb = f"{(len(contents) / (1024 * 1024)):.2f} MB"
    file_type = file.filename.split(".")[-1].lower() if "." in file.filename else "txt"
    
    # 1. Create document metadata record in database
    doc_id = f"doc-{int(time.time() * 1000)}"
    new_doc = Document(
        id=doc_id,
        chat_id=chat_id,
        user_id=current_user.id,
        file_name=file.filename,
        file_size=file_size_mb,
        file_type=file_type,
        status="parsing",
        chunks_count=0
    )
    db.add(new_doc)
    db.commit()
    db.refresh(new_doc)

    # 2. Extract text from the raw binary contents
    text_content = ""
    try:
        if file_type == "txt" or file_type == "md":
            text_content = contents.decode("utf-8", errors="ignore")
        elif file_type == "pdf":
            import io
            import pypdf
            pdf_file = io.BytesIO(contents)
            reader = pypdf.PdfReader(pdf_file)
            text_content = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
        elif file_type == "docx":
            import io
            import docx2txt
            docx_file = io.BytesIO(contents)
            text_content = docx2txt.process(docx_file)
        elif file_type == "csv":
            text_content = contents.decode("utf-8", errors="ignore")
        elif file_type == "xlsx":
            # For excel sheets, join all sheet values as space-separated tokens
            import io
            import openpyxl
            excel_file = io.BytesIO(contents)
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            text_rows = []
            for sheet_name in wb.sheetnames[:2]:
                sheet = wb[sheet_name]
                for r in sheet.iter_rows(values_only=True):
                    text_rows.append(" ".join([str(v) for v in r if v is not None]))
            text_content = "\n".join(text_rows)
        else:
            # Fallback for unrecognized formats (images/media)
            text_content = f"Uploaded attachment: {file.filename} of format {file_type}."
    except Exception as e:
        new_doc.status = "error"
        db.commit()
        raise HTTPException(status_code=400, detail=f"Failed to extract document contents: {e}")

    # 3. Index contents into RAG database
    try:
        chunks_count = rag_pipeline.ingest_document(
            doc_id=doc_id,
            doc_name=file.filename,
            content=text_content
        )
        new_doc.chunks_count = chunks_count
        new_doc.status = "ready"
        db.commit()
    except Exception as e:
        new_doc.status = "error"
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed indexing text vector chunks: {e}")

    return {
        "id": doc_id,
        "name": file.filename,
        "size": file_size_mb,
        "type": file_type,
        "status": "ready",
        "progress": 100,
        "uploadedAt": datetime.now().strftime("%I:%M %p")
    }

@router.delete("/{doc_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_document(doc_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(Document.id == doc_id, Document.user_id == current_user.id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Optional: Delete chunks from vector database if needed
    db.delete(doc)
    db.commit()
    return
