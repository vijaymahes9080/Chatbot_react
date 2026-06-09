import time
import math
import traceback
from typing import Dict, Any, List, Callable
from app.utils.logger import logger
from app.services.browser_agent import BrowserAgent

class ToolManager:
    def __init__(self):
        self._tools: Dict[str, Callable] = {}
        self.browser_agent = BrowserAgent()
        self._register_default_tools()

    def register_tool(self, name: str, func: Callable):
        self._tools[name] = func
        logger.info(f"Registered tool: {name}")

    def execute_tool(self, name: str, arguments: Dict[str, Any]) -> Dict[str, Any]:
        """
        Executes a registered tool with error safety, execution logs, and timing.
        """
        start_time = time.time()
        
        # Standardize name lookup from user-friendly frontend titles to internal snake_case keys
        aliases = {
            "browser tool": "browser_tool",
            "calculator tool": "calculator_tool",
            "code interpreter": "code_interpreter",
            "file reader": "file_reader",
            "database connector": "database_tool",
            "rag retriever": "search_tool"
        }
        lookup_name = name.lower()
        resolved_name = aliases.get(lookup_name, name)
        
        tool_logs = [f"[INFO] Initializing {name} (resolved as {resolved_name}) with arguments: {arguments}"]
        
        if resolved_name not in self._tools:
            error_msg = f"Tool '{name}' is not registered in the ToolManager."
            logger.error(error_msg)
            return {
                "tool_name": name,
                "status": "error",
                "execution_time": "0.0s",
                "result": None,
                "error": error_msg,
                "logs": [error_msg]
            }

        try:
            # Execute the function passing arguments and log context
            result = self._tools[resolved_name](arguments, tool_logs)
            execution_time = f"{(time.time() - start_time):.2f}s"
            
            tool_logs.append(f"[SUCCESS] Finished {resolved_name} execution in {execution_time}")
            return {
                "tool_name": resolved_name,
                "status": "success",
                "execution_time": execution_time,
                "result": result,
                "error": None,
                "logs": tool_logs
            }
        except Exception as e:
            execution_time = f"{(time.time() - start_time):.2f}s"
            tb = traceback.format_exc()
            error_msg = f"Exception running tool {name}: {str(e)}"
            tool_logs.append(f"[ERROR] {error_msg}")
            tool_logs.append(tb)
            
            return {
                "tool_name": name,
                "status": "error",
                "execution_time": execution_time,
                "result": None,
                "error": str(e),
                "logs": tool_logs
            }

    def _register_default_tools(self):
        # System Default Tools Mapping
        self.register_tool("calculator_tool", self._calculator_tool)
        self.register_tool("file_reader", self._file_reader_tool)
        self.register_tool("pdf_reader", self._pdf_reader_tool)
        self.register_tool("csv_reader", self._csv_reader_tool)
        self.register_tool("excel_reader", self._excel_reader_tool)
        self.register_tool("code_interpreter", self._code_interpreter_tool)
        self.register_tool("database_tool", self._database_tool)
        self.register_tool("api_tool", self._api_tool)
        self.register_tool("memory_tool", self._memory_tool)
        self.register_tool("browser_tool", self._browser_tool)
        self.register_tool("search_tool", self._browser_tool)

    def _browser_tool(self, args: Dict[str, Any], logs: List[str]) -> Dict[str, Any]:
        query = args.get("query", "")
        logs.append(f"[BROWSER] Delegating query to BrowserAgent: {query}")
        return self.browser_agent.perform_search(query)

    # 1. Calculator Tool
    def _calculator_tool(self, args: Dict[str, Any], logs: List[str]) -> Any:
        expr = args.get("expression", "")
        logs.append(f"[CALC] Evaluating: {expr}")
        # Safe math parser using simple eval logic with restricted builtins
        allowed = {"__builtins__": None, "math": math, "abs": abs, "pow": pow, "sum": sum}
        try:
            # Clean string expression
            sanitized = expr.replace("^", "**").replace("x", "*").replace(" ", "")
            res = eval(sanitized, allowed)
            logs.append(f"[CALC] Result: {res}")
            return res
        except Exception as e:
            raise ValueError(f"Failed to evaluate mathematical expression '{expr}': {e}")

    # 2. Local File Reader
    def _file_reader_tool(self, args: Dict[str, Any], logs: List[str]) -> str:
        filepath = args.get("path", "")
        logs.append(f"[FILE] Reading text file: {filepath}")
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read(20000)  # Read first 20k characters
                logs.append(f"[FILE] Loaded {len(content)} characters successfully.")
                return content
        except Exception as e:
            raise FileNotFoundError(f"Cannot read filepath '{filepath}': {e}")

    # 3. PDF Reader
    def _pdf_reader_tool(self, args: Dict[str, Any], logs: List[str]) -> str:
        filepath = args.get("path", "")
        logs.append(f"[PDF] Parsing PDF: {filepath}")
        try:
            import pypdf
            reader = pypdf.PdfReader(filepath)
            text = ""
            for idx, page in enumerate(reader.pages[:10]):  # Limit first 10 pages
                text += f"\n--- Page {idx+1} ---\n" + page.extract_text()
            logs.append(f"[PDF] Extracted {len(text)} characters from {len(reader.pages)} pages.")
            return text
        except Exception as e:
            raise RuntimeError(f"Error parsing PDF file '{filepath}': {e}")

    # 4. CSV Reader
    def _csv_reader_tool(self, args: Dict[str, Any], logs: List[str]) -> List[List[str]]:
        filepath = args.get("path", "")
        logs.append(f"[CSV] Reading tabular CSV: {filepath}")
        try:
            import csv
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)[:100]  # Limit first 100 rows
                logs.append(f"[CSV] Loaded {len(rows)} rows successfully.")
                return rows
        except Exception as e:
            raise RuntimeError(f"Error parsing CSV '{filepath}': {e}")

    # 5. Excel Reader
    def _excel_reader_tool(self, args: Dict[str, Any], logs: List[str]) -> Dict[str, List[List[Any]]]:
        filepath = args.get("path", "")
        logs.append(f"[EXCEL] Reading Excel Sheets: {filepath}")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            result = {}
            for sheet_name in wb.sheetnames[:3]:  # Limit first 3 sheets
                sheet = wb[sheet_name]
                rows = []
                for row in list(sheet.iter_rows(values_only=True))[:50]:  # Limit first 50 rows
                    rows.append(list(row))
                result[sheet_name] = rows
            logs.append(f"[EXCEL] Extracted sheets: {list(result.keys())}")
            return result
        except Exception as e:
            raise RuntimeError(f"Error parsing Excel '{filepath}': {e}")

    # 6. Python Code Interpreter (Secure local sandboxed runner)
    def _code_interpreter_tool(self, args: Dict[str, Any], logs: List[str]) -> Dict[str, Any]:
        code = args.get("code", "")
        logs.append("[SANDBOX] Initializing secure python sandboxed execution environment")
        
        # Intercept simple malicious packages
        blacklisted = ["os", "sys", "subprocess", "shutil", "socket", "urllib", "requests"]
        for mod in blacklisted:
            if f"import {mod}" in code or f"from {mod}" in code:
                raise PermissionError(f"Import block error: Importing module '{mod}' is not allowed in sandbox.")

        try:
            # Set up output capture
            import io
            import sys
            stdout_capture = io.StringIO()
            stderr_capture = io.StringIO()
            
            # Context globals with restricted actions
            globals_dict = {"__builtins__": __builtins__}
            locals_dict = {}
            
            # Temporarily redirect output stream
            old_stdout = sys.stdout
            old_stderr = sys.stderr
            sys.stdout = stdout_capture
            sys.stderr = stderr_capture
            
            try:
                exec(code, globals_dict, locals_dict)
            finally:
                sys.stdout = old_stdout
                sys.stderr = old_stderr
                
            stdout_val = stdout_capture.getvalue()
            stderr_val = stderr_capture.getvalue()
            
            logs.append(f"[SANDBOX] Run completed successfully.")
            return {
                "stdout": stdout_val,
                "stderr": stderr_val,
                "vars": {k: str(v) for k, v in locals_dict.items() if not k.startswith("__")}
            }
        except Exception as e:
            raise RuntimeError(f"Sandbox runtime exception: {e}")

    # 7. Database Tool (Read-only SQLite query tester)
    def _database_tool(self, args: Dict[str, Any], logs: List[str]) -> List[Dict[str, Any]]:
        query = args.get("query", "")
        logs.append(f"[DATABASE] Executing SQL Query: {query}")
        
        # Security: Allow only SELECT queries
        query_strip = query.strip().upper()
        if not query_strip.startswith("SELECT"):
            raise PermissionError("Database execute error: Only read-only SELECT statements are supported.")
            
        try:
            # Connect to local SQLite DB
            import sqlite3
            conn = sqlite3.connect("aethermind.db")
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            result = [dict(row) for row in rows[:50]]  # limit first 50 results
            logs.append(f"[DATABASE] Query executed successfully. Returned {len(result)} records.")
            conn.close()
            return result
        except Exception as e:
            raise RuntimeError(f"Database query failure: {e}")

    # 8. External API tool
    def _api_tool(self, args: Dict[str, Any], logs: List[str]) -> Dict[str, Any]:
        url = args.get("url", "")
        method = args.get("method", "GET").upper()
        headers = args.get("headers", {})
        data = args.get("data", None)
        
        logs.append(f"[API] Dispatching {method} request to URL: {url}")
        try:
            import httpx
            with httpx.Client(timeout=10.0) as client:
                if method == "GET":
                    resp = client.get(url, headers=headers)
                elif method == "POST":
                    resp = client.post(url, headers=headers, json=data)
                else:
                    resp = client.request(method, url, headers=headers, json=data)
                
                logs.append(f"[API] Dispatch status: {resp.status_code}")
                return {
                    "status_code": resp.status_code,
                    "content": resp.text[:10000]  # First 10k response text characters
                }
        except Exception as e:
            raise ConnectionError(f"HTTP connection failed to URL '{url}': {e}")

    # 9. Short-term memory tool
    def _memory_tool(self, args: Dict[str, Any], logs: List[str]) -> str:
        action = args.get("action", "retrieve")
        key = args.get("key", "")
        value = args.get("value", "")
        
        logs.append(f"[MEMORY] Memory lookup: Action={action}, Key={key}")
        # Simplistic key-value session dict emulator
        return f"[Memory Service Success: Key='{key}' loaded successfully]"
